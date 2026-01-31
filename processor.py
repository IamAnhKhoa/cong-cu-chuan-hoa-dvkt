# -*- coding: utf-8 -*-
"""
Data Processing Module for Technical Catalog Standardization
Handles Excel file reading, matching, and transformation
"""

import pandas as pd
from fuzzywuzzy import fuzz, process
from typing import Dict, List, Tuple, Optional, Callable
import os
from logger_config import AppLogger


class CatalogProcessor:
    """Process technical catalog data with fuzzy matching"""
    
    def __init__(self, logger: AppLogger = None):
        self.ref_quy_trinh_df = None
        self.ref_gia_hdnd_df = None
        self.ref_dvkt_gia_max_df = None
        self.reference_folder = None
        self.logger = logger
        
    def load_reference_files(self, folder_path: str) -> Tuple[bool, str]:
        """Load reference Excel files from folder"""
        try:
            quy_trinh_path = os.path.join(folder_path, "QUY_TRINH_DVKT_BYT.xlsx")
            gia_hdnd_path = os.path.join(folder_path, "GIA_HDND.xlsx")
            max_gia_path = os.path.join(folder_path, "DVKT_GIA_MAX.xlsx")
            
            if not os.path.exists(quy_trinh_path):
                return False, f"Không tìm thấy file: QUY_TRINH_DVKT_BYT.xlsx"
            
            if not os.path.exists(gia_hdnd_path):
                return False, f"Không tìm thấy file: GIA_HDND.xlsx"
            
            # Helper to load and find correct sheet
            def load_with_correct_sheet(path, expected_col_keyword):
                xl = pd.ExcelFile(path, engine='openpyxl')
                target_sheet = None
                
                # Check known sheet names first
                if 'QUY TRINH' in xl.sheet_names:
                    target_sheet = 'QUY TRINH'
                
                # If not found or just to be safe, search for content
                if target_sheet is None:
                    for sheet in xl.sheet_names:
                        df = pd.read_excel(path, sheet_name=sheet, nrows=5, engine='openpyxl')
                        # Check if any column contains the keyword
                        col_match = any(expected_col_keyword.lower() in str(col).lower() for col in df.columns)
                        if col_match:
                            target_sheet = sheet
                            break
                            
                # Fallback to first sheet if nothing found (will likely fail later but better than crash)
                if target_sheet is None:
                    target_sheet = xl.sheet_names[0]
                    
                return pd.read_excel(path, sheet_name=target_sheet, engine='openpyxl')

            # Load files with smart sheet detection
            self.ref_quy_trinh_df = load_with_correct_sheet(quy_trinh_path, 'dịch vụ')
            self.ref_gia_hdnd_df = pd.read_excel(gia_hdnd_path, engine='openpyxl')
            
            # Optional third file
            if os.path.exists(max_gia_path):
                 self.ref_dvkt_gia_max_df = pd.read_excel(max_gia_path, engine='openpyxl')
                 max_msg = " và DVKT_GIA_MAX"
            else:
                 self.ref_dvkt_gia_max_df = None
                 max_msg = ""
            
            self.reference_folder = folder_path
            
            return True, f"Đã tải {len(self.ref_quy_trinh_df)} dịch vụ từ QUY_TRINH, {len(self.ref_gia_hdnd_df)} dịch vụ từ GIA_HDND{max_msg}"
            
        except Exception as e:
            return False, f"Lỗi khi tải file tham chiếu: {str(e)}"
    
    def find_best_match(self, service_name: str, reference_names: List[str], threshold: int = 80) -> Optional[Tuple[str, int]]:
        """Find best matching service name using exact match first, then fuzzy matching
        
        Returns:
            - (matched_name, score): Normal match found
            - ('AMBIGUOUS', list_of_matches): Multiple similar matches found (score >= 95%)
            - None: No match found
        """
        if not service_name or not reference_names:
            return None
        
        # Normalize for comparison
        service_name_normalized = service_name.strip().lower()
        
        # Check for multiple exact matches with same base name
        exact_matches = []
        service_name_clean = service_name_normalized.split('[')[0].strip()
        
        for ref_name in reference_names:
            ref_name_clean = ref_name.strip().lower().split('[')[0].strip()
            if ref_name_clean == service_name_clean:
                exact_matches.append(ref_name)
        
        # If multiple exact base matches, it's ambiguous
        if len(exact_matches) > 1:
            return ('AMBIGUOUS', exact_matches)
        
        # If single exact match, return it
        if len(exact_matches) == 1:
            return exact_matches[0], 100
        
        # Priority: Full exact match (case-insensitive)
        for ref_name in reference_names:
            if ref_name.strip().lower() == service_name_normalized:
                return ref_name, 100  # Exact match = 100% score
        
        # Fuzzy matching with ambiguity detection
        all_results = process.extract(service_name, reference_names, scorer=fuzz.token_sort_ratio, limit=5)
        
        if not all_results:
            return None
        
        # Get top results with score >= 95 (very similar)
        high_score_matches = [(name, score) for name, score in all_results if score >= 95]
        
        # If multiple high-score matches, it's ambiguous
        if len(high_score_matches) > 1:
            match_names = [name for name, score in high_score_matches]
            return ('AMBIGUOUS', match_names)
        
        # Single best match
        best_match = all_results[0]
        if best_match[1] >= threshold:
            return best_match[0], best_match[1]
        
        return None
    
    def find_best_match_with_chapter(self, service_name: str, chapter_name: str, 
                                     reference_df: pd.DataFrame, ref_name_col: str, 
                                     ref_chapter_col: str, threshold: int = 80) -> Optional[Tuple[str, int]]:
        """Find best match with chapter filtering for higher accuracy
        
        Args:
            service_name: Service name to match
            chapter_name: Chapter name from input
            reference_df: Reference dataframe
            ref_name_col: Column name for service names in reference
            ref_chapter_col: Column name for chapter in reference (can be None)
            threshold: Matching threshold
            
        Returns:
            Same as find_best_match: (matched_name, score), ('AMBIGUOUS', list), or None
        """
        if ref_chapter_col and chapter_name and not pd.isna(chapter_name):
            # Filter by chapter first
            chapter_name_normalized = str(chapter_name).strip().lower()
            
            # Find matching chapter rows
            filtered_df = reference_df[
                reference_df[ref_chapter_col].astype(str).str.strip().str.lower() == chapter_name_normalized
            ]
            
            if len(filtered_df) > 0:
                # Match within same chapter
                filtered_names = filtered_df[ref_name_col].dropna().astype(str).str.strip().tolist()
                return self.find_best_match(service_name, filtered_names, threshold)
        
        # Fallback to full match if no chapter info or no chapter matches
        all_names = reference_df[ref_name_col].dropna().astype(str).str.strip().tolist()
        return self.find_best_match(service_name, all_names, threshold)
    
    def process_quy_trinh_file(self, input_file: str, output_file: str, threshold: int = 80, progress_callback: Callable[[str], None] = None) -> Tuple[bool, str, Dict]:
        """
        Process file 1: QUY_TRINH_DVKT_BYT → DM_DVKT_TT21_cong
        
        Input columns: Must contain "Tên kỹ thuật"
        Output columns: STT, MA_DICH_VU, TEN_DICH_VU, DON_GIA, QUY_TRINH, CSKCB_CGKT, CSKCB_CLS
        """
        try:
            # Log processing start
            if self.logger:
                self.logger.log_processing_start(input_file, "Loại 1: QUY_TRINH → DM_DVKT_TT21_cong")
            
            # Validate reference data loaded
            if self.ref_quy_trinh_df is None:
                return False, "Chưa tải file tham chiếu. Vui lòng chọn thư mục chứa file gốc.", {}
            
            # Read input file with proper encoding
            input_df = pd.read_excel(input_file, engine='openpyxl')
            
            # Check for required column
            name_column = None
            input_df.columns = input_df.columns.str.strip()  # Normalize headers
            
            possible_headers = ['tên kỹ thuật', 'tên dịch vụ', 'tên dvkt', 'ten dich vu', 'ten ky thuat', 'tên']
            
            # Try exact match first
            for col in input_df.columns:
                if col.lower() in possible_headers:
                    name_column = col
                    break
            
            # Fuzzy header match if exact match fails
            if name_column is None:
                for col in input_df.columns:
                    col_lower = col.lower()
                    if 'tên' in col_lower and ('thuật' in col_lower or 'dịch vụ' in col_lower or 'dvkt' in col_lower):
                        name_column = col
                        break
            
            if name_column is None:
                return False, f'Không tìm thấy cột Tên dịch vụ/kỹ thuật. Các cột có trong file: {", ".join(input_df.columns)}', {}
            
            # Prepare reference names
            ref_name_column = None
            self.ref_quy_trinh_df.columns = self.ref_quy_trinh_df.columns.str.strip()
            
            for col in self.ref_quy_trinh_df.columns:
                if 'TEN_DICH_VU' in col or ('tên' in col.lower() and 'vụ' in col.lower()):
                    ref_name_column = col
                    break
                    
            if ref_name_column is None:
                return False, "Không tìm thấy cột tên dịch vụ trong file QUY_TRINH_DVKT_BYT.xlsx", {}
            
            # Find MA_DICH_VU column in reference file
            ref_code_column = None
            for col in self.ref_quy_trinh_df.columns:
                if 'MA_DICH_VU' in col or ('mã' in col.lower() and 'vụ' in col.lower()):
                    ref_code_column = col
                    break
            
            reference_names = self.ref_quy_trinh_df[ref_name_column].dropna().astype(str).str.strip().tolist()
            
            # Find MA_DICH_VU column in input file
            input_code_column = None
            for col in input_df.columns:
                if 'MA_DICH_VU' in col or ('mã' in col.lower() and ('vụ' in col.lower() or 'dịch' in col.lower())):
                    input_code_column = col
                    break
            
            # Process matching
            results = []
            unmatched_records = []  # Store unmatched/ambiguous records
            stats = {
                'total': len(input_df),
                'matched': 0,
                'matched_by_code': 0,
                'matched_by_name': 0,
                'unmatched': 0,
                'ambiguous': 0,
                'match_details': []
            }
            
            for idx, row in input_df.iterrows():
                raw_name = row[name_column]
                # Skip if empty or NaN
                if pd.isna(raw_name) or str(raw_name).strip() == "":
                    continue
                    
                service_name = str(raw_name).strip()
                matched_by_code = False
                match_result = None
                
                # PRIORITY 1: Try exact match by CODE if available
                if input_code_column and ref_code_column:
                    input_code = row.get(input_code_column)
                    if not pd.isna(input_code) and str(input_code).strip():
                        code_to_match = str(input_code).strip()
                        # Exact match by code
                        code_matches = self.ref_quy_trinh_df[
                            self.ref_quy_trinh_df[ref_code_column].astype(str).str.strip() == code_to_match
                        ]
                        if len(code_matches) > 0:
                            # Found exact code match
                            matched_by_code = True
                            match_result = (code_matches.iloc[0][ref_name_column], 100)
                
                # PRIORITY 2: Fallback to fuzzy name matching if no code match
                if not matched_by_code:
                    match_result = self.find_best_match(service_name, reference_names, threshold)
                
                # Check for ambiguous match
                if match_result and match_result[0] == 'AMBIGUOUS':
                    # Multiple similar matches - use first one but warn user
                    ambiguous_list = match_result[1]
                    first_match = ambiguous_list[0]
                    
                    # Get data from first match
                    ref_matches = self.ref_quy_trinh_df[self.ref_quy_trinh_df[ref_name_column] == first_match]
                    if len(ref_matches) > 0:
                        ref_row = ref_matches.iloc[0]
                        
                        # Create warning text for last column
                        warning = f"⚠️ CÓ {len(ambiguous_list)} DỊCH VỤ TƯƠNG TỰ - XEM KỸ"
                        
                        output_row = {
                            'STT': idx + 1,
                            'MA_DICH_VU': ref_row.get('MA_DICH_VU', ''),
                            'TEN_DICH_VU': ref_row.get(ref_name_column, first_match),
                            'DON_GIA': ref_row.get('DON_GIA', ''),
                            'QUY_TRINH': ref_row.get('QUY_TRINH', ''),
                            'CSKCB_CGKT': '',
                            'CSKCB_CLS': warning  # Warning in last column
                        }
                        results.append(output_row)
                        stats['ambiguous'] += 1
                        stats['match_details'].append({
                            'input': service_name,
                            'matched': f'{first_match} (⚠️ {len(ambiguous_list)} tương tự)',
                            'score': 99
                        })
                        
                        # Still add to unmatched file with full details
                        warning_text = f'⚠️ CÓ {len(ambiguous_list)} DỊCH VỤ TƯƠNG TỰ:\n'
                        for i, match in enumerate(ambiguous_list[:5], 1):
                            warning_text += f'{i}. {match}\n'
                        if len(ambiguous_list) > 5:
                            warning_text += f'... và {len(ambiguous_list) - 5} dịch vụ khác'
                        
                        unmatched_records.append({
                            'STT': idx + 1,
                            'TEN_DICH_VU_GOC': service_name,
                            'GHI_CHU': warning_text
                        })
                        
                        if self.logger:
                            self.logger.logger.warning(f'Ambiguous match for "{service_name}": Using first of {len(ambiguous_list)} options')
                        if progress_callback:
                            progress_callback(f"Đã xử lý {idx + 1}/{len(input_df)}: Cảnh báo - {service_name[:50]}...")
                    else:
                        # Fallback if first match not found
                        output_row = {
                            'STT': idx + 1,
                            'MA_DICH_VU': '',
                            'TEN_DICH_VU': service_name,
                            'DON_GIA': '',
                            'QUY_TRINH': '',
                            'CSKCB_CGKT': '',
                            'CSKCB_CLS': f'⚠️ CÓ {len(ambiguous_list)} DỊCH VỤ - KIỂM TRA'
                        }
                        results.append(output_row)
                        stats['ambiguous'] += 1
                
                elif match_result:
                    # Normal single match
                    matched_name, score = match_result
                    ref_matches = self.ref_quy_trinh_df[self.ref_quy_trinh_df[ref_name_column] == matched_name]
                    
                    if len(ref_matches) == 0:
                        # Matched name not found in reference (shouldn't happen but safeguard)
                        output_row = {
                            'STT': idx + 1,
                            'MA_DICH_VU': '',
                            'TEN_DICH_VU': service_name,
                            'DON_GIA': '',
                            'QUY_TRINH': '',
                            'CSKCB_CGKT': '',
                            'CSKCB_CLS': ''
                        }
                        results.append(output_row)
                        stats['unmatched'] += 1
                        if self.logger:
                            self.logger.logger.error(f'Match "{matched_name}" not found in reference for "{service_name}"')
                        continue
                    
                    ref_row = ref_matches.iloc[0]
                    
                    output_row = {
                        'STT': idx + 1,
                        'MA_DICH_VU': ref_row.get('MA_DICH_VU', ''),
                        'TEN_DICH_VU': ref_row.get(ref_name_column, matched_name),
                        'DON_GIA': ref_row.get('DON_GIA', ''),
                        'QUY_TRINH': ref_row.get('QUY_TRINH', ''),
                        'CSKCB_CGKT': '',
                        'CSKCB_CLS': ''
                    }
                    results.append(output_row)
                    stats['matched'] += 1
                    if matched_by_code:
                        stats['matched_by_code'] += 1
                    else:
                        stats['matched_by_name'] += 1
                    stats['match_details'].append({
                        'input': service_name,
                        'matched': matched_name,
                        'score': score
                    })
                    
                    if self.logger:
                        self.logger.log_match(service_name, matched_name, score)
                    if progress_callback:
                        progress_callback(f"Đã xử lý {idx + 1}/{len(input_df)}: Khớp - {service_name[:50]}...")
                
                else:
                    # No match found
                    output_row = {
                        'STT': idx + 1,
                        'MA_DICH_VU': '',
                        'TEN_DICH_VU': service_name,
                        'DON_GIA': '',
                        'QUY_TRINH': '',
                        'CSKCB_CGKT': '',
                        'CSKCB_CLS': ''
                    }
                    results.append(output_row)
                    stats['unmatched'] += 1
                    stats['match_details'].append({
                        'input': service_name,
                        'matched': 'KHÔNG TÌM THẤY',
                        'score': 0
                    })
                    
                    # Store unmatched record with notes
                    unmatched_records.append({
                        'STT': idx + 1,
                        'TEN_DICH_VU_GOC': service_name,
                        'GHI_CHU': f'Không tìm thấy tên tương ứng trong file gốc (Ngưỡng khớp: {threshold}%). Vui lòng kiểm tra lại hoặc thêm thủ công.'
                    })
                    
                    # Log no match
                    if self.logger:
                        self.logger.log_no_match(service_name, threshold)
                    
                    # Progress callback
                    if progress_callback:
                        progress_callback(f"Đã xử lý {idx + 1}/{len(input_df)}: Không khớp - {service_name[:50]}...")
            
            # Create output DataFrame
            output_df = pd.DataFrame(results)
            
            # Save to Excel with proper encoding
            output_df.to_excel(output_file, index=False, engine='openpyxl')
            
            # Apply yellow highlighting to ambiguous rows
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            import os
            
            wb = load_workbook(output_file)
            ws = wb.active
            
            # Yellow fill for ambiguous warnings
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # Find rows with ambiguous warnings (check last column for "⚠️")
            last_col_idx = ws.max_column
            for row_idx in range(2, ws.max_row + 1):  # Start from 2 to skip header
                last_cell = ws.cell(row=row_idx, column=last_col_idx)
                if last_cell.value and isinstance(last_cell.value, str) and '⚠️' in last_cell.value:
                    # Highlight entire row in yellow
                    for col_idx in range(1, last_col_idx + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
            
            # Format Excel for better readability
            # 1. Bold headers
            for col_idx in range(1, last_col_idx + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 2. Auto-fit column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                adjusted_width = min(length + 2, 50)  # Max width 50
                ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
            # 3. Wrap text for long cells
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            wb.save(output_file)
            
            # Auto-open Excel file
            try:
                os.startfile(output_file)  # Windows only
            except:
                pass  # Silently fail if not Windows or file can't be opened
            
            # Export unmatched records if any
            if unmatched_records:
                base_name = os.path.splitext(output_file)[0]
                unmatched_file = f"{base_name}_unmatched.xlsx"
                unmatched_df = pd.DataFrame(unmatched_records)
                unmatched_df.to_excel(unmatched_file, index=False, engine='openpyxl')
                stats['unmatched_file'] = unmatched_file
                if self.logger:
                    self.logger.logger.info(f"Đã xuất {len(unmatched_records)} bản ghi không khớp ra: {unmatched_file}")
            
            # Log processing end
            if self.logger:
                self.logger.log_processing_end(stats)
            
            message = f"Xử lý thành công!\n"
            message += f"Tổng số: {stats['total']}\n"
            message += f"Khớp: {stats['matched']}\n"
            message += f"Không khớp: {stats['unmatched']}"
            if stats.get('unmatched_file'):
                message += f"\n\nĐã xuất các bản ghi không khớp ra file:\n{stats['unmatched_file']}"
            
            return True, message, stats
            
        except Exception as e:
            if self.logger:
                self.logger.log_error(str(e))
            return False, f"Lỗi khi xử lý file: {str(e)}", {}
    
    def process_gia_hdnd_file(self, input_file: str, output_file: str, threshold: int = 80, progress_callback: Callable[[str], None] = None) -> Tuple[bool, str, Dict]:
        """
        Process file 2: GIA_HDND → Standardized output
        
        Output columns: STT, MA_TUONG_DUONG, TEN_DVKT_PHEDUYET, TEN_DVKT_GIA, PHAN_LOAI_PTTT,
                       DON_GIA, GHI_CHU, QUYET_DINH, QUY_TRINH, TU_NGAY, DEN_NGAY, CSKCB_CGKT, CSKCB_CLS
        """
        try:
            # Log processing start
            if self.logger:
                self.logger.log_processing_start(input_file, "Loại 2: GIA_HDND → File chuẩn")
            
            # Validate reference data loaded
            if self.ref_gia_hdnd_df is None or self.ref_quy_trinh_df is None:
                return False, "Chưa tải đầy đủ file tham chiếu. Vui lòng chọn thư mục chứa file gốc.", {}
            
            # Read input file with proper encoding
            input_df = pd.read_excel(input_file, engine='openpyxl')
            
            # Check for required columns
            name_column = None
            chapter_column = None  # New: Tên chương column
            input_df.columns = input_df.columns.str.strip()  # Normalize headers
            
            # Find service name column
            possible_headers = ['tên kỹ thuật', 'tên dịch vụ', 'tên dvkt', 'ten dich vu', 'ten ky thuat', 'tên']
            
            for col in input_df.columns:
                if col.lower() in possible_headers:
                    name_column = col
                    break
            
            if name_column is None:
                for col in input_df.columns:
                     col_lower = col.lower()
                     if 'tên' in col_lower and ('thuật' in col_lower or 'dịch vụ' in col_lower or 'dvkt' in col_lower):
                        name_column = col
                        break
            
            if name_column is None:
                return False, f'Không tìm thấy cột Tên dịch vụ/kỹ thuật. Các cột có trong file: {", ".join(input_df.columns)}', {}
            
            # Find chapter column (Tên chương)
            possible_chapter_headers = ['tên chương', 'ten chuong', 'chương']
            for col in input_df.columns:
                if col.lower() in possible_chapter_headers:
                    chapter_column = col
                    break
            
            if chapter_column is None:
                for col in input_df.columns:
                    col_lower = col.lower()
                    if 'chương' in col_lower or 'chuong' in col_lower:
                        chapter_column = col
                        break
            
            # Prepare reference names from GIA_HDND
            ref_name_column = None
            ref_chapter_column = None  # New: Chapter column in reference
            self.ref_gia_hdnd_df.columns = self.ref_gia_hdnd_df.columns.str.strip()
            
            for col in self.ref_gia_hdnd_df.columns:
                if 'Tên dịch vụ kỹ thuật' in col or ('TT23' in col and 'Tên' in col):
                    ref_name_column = col
                    break
            
            # Find chapter column in reference
            for col in self.ref_gia_hdnd_df.columns:
                col_lower = col.lower()
                if 'tên chương' in col_lower or 'chương' in col_lower or 'chuong' in col_lower:
                    ref_chapter_column = col
                    break
                    
            if ref_name_column is None:
                return False, "Không tìm thấy cột tên dịch vụ trong file GIA_HDND.xlsx", {}
            
            reference_names = self.ref_gia_hdnd_df[ref_name_column].dropna().astype(str).str.strip().tolist()

            # Find MA_TUONG_DUONG column in GIA_HDND reference
            ref_code_column_gia = None
            for col in self.ref_gia_hdnd_df.columns:
                if 'MA_TUONG_DUONG' in col or 'MA_DICH_VU' in col or ('mã' in col.lower() and ('tương' in col.lower() or 'vụ' in col.lower())):
                    ref_code_column_gia = col
                    break
            
            # Prepare reference names from DVKT_GIA_MAX if available
            max_reference_names = []
            max_ref_name_col = None
            max_ref_code_col = None
            if self.ref_dvkt_gia_max_df is not None:
                self.ref_dvkt_gia_max_df.columns = self.ref_dvkt_gia_max_df.columns.str.strip()
                for col in self.ref_dvkt_gia_max_df.columns:
                    if 'TEN_DVKT_GIA' in col or ('Tên' in col and 'Gia' in col):
                        max_ref_name_col = col
                        break
                # Find code column in DVKT_GIA_MAX
                for col in self.ref_dvkt_gia_max_df.columns:
                    if 'MA' in col and ('DICH_VU' in col or 'DVKT' in col or 'TUONG_DUONG' in col):
                        max_ref_code_col = col
                        break
                if max_ref_name_col:
                    max_reference_names = self.ref_dvkt_gia_max_df[max_ref_name_col].dropna().astype(str).str.strip().tolist()
            
            # Find MA_DICH_VU or MA_TUONG_DUONG column in input file
            input_code_column = None
            for col in input_df.columns:
                if 'MA_DICH_VU' in col or 'MA_TUONG_DUONG' in col or ('mã' in col.lower() and ('vụ' in col.lower() or 'dịch' in col.lower() or 'tương' in col.lower())):
                    input_code_column = col
                    break

            
            # Also prepare QUY_TRINH reference for combining
            quy_trinh_name_col = None
            self.ref_quy_trinh_df.columns = self.ref_quy_trinh_df.columns.str.strip()
            
            for col in self.ref_quy_trinh_df.columns:
                if 'TEN_DICH_VU' in col or ('tên' in col.lower() and 'vụ' in col.lower()):
                    quy_trinh_name_col = col
                    break
            
            # Process matching
            results = []
            unmatched_records = []  # Store unmatched records
            stats = {
                'total': len(input_df),
                'matched': 0,
                'unmatched': 0,
                'ambiguous': 0,
                'match_details': [],
                'unmatched_file': None
            }
            
            for idx, row in input_df.iterrows():
                raw_name = row[name_column]
                if pd.isna(raw_name) or str(raw_name).strip() == "":
                    continue
                service_name = str(raw_name).strip()
                
                # Get chapter name if available
                chapter_name = None
                if chapter_column and chapter_column in row:
                    chapter_name = row[chapter_column] if not pd.isna(row[chapter_column]) else None
                
                # Initialize matching flags
                matched_by_code_gia = False
                matched_by_code_max = False
                match_result_gia = None
                match_result_max = None
                
                # PRIORITY 1: Try exact match by CODE if available
                if input_code_column:
                    input_code = row.get(input_code_column)
                    if not pd.isna(input_code) and str(input_code).strip():
                        code_to_match = str(input_code).strip()
                        
                        # Try to match in GIA_HDND by code
                        if ref_code_column_gia:
                            gia_code_matches = self.ref_gia_hdnd_df[
                                self.ref_gia_hdnd_df[ref_code_column_gia].astype(str).str.strip() == code_to_match
                            ]
                            if len(gia_code_matches) > 0:
                                matched_by_code_gia = True
                                match_result_gia = (gia_code_matches.iloc[0][ref_name_column], 100)
                        
                        # Try to match in DVKT_GIA_MAX by code
                        if max_ref_code_col and self.ref_dvkt_gia_max_df is not None:
                            max_code_matches = self.ref_dvkt_gia_max_df[
                                self.ref_dvkt_gia_max_df[max_ref_code_col].astype(str).str.strip() == code_to_match
                            ]
                            if len(max_code_matches) > 0:
                                matched_by_code_max = True
                                match_result_max = (max_code_matches.iloc[0][max_ref_name_col], 100)
                
                # PRIORITY 2: Fallback to fuzzy name matching if no code match
                # 1. Find best match in GIA_HDND (with chapter filtering if available)
                if not matched_by_code_gia:
                    if ref_chapter_column and chapter_name:
                        match_result_gia = self.find_best_match_with_chapter(
                            service_name, chapter_name, self.ref_gia_hdnd_df, 
                            ref_name_column, ref_chapter_column, threshold)
                    else:
                        match_result_gia = self.find_best_match(service_name, reference_names, threshold)
                
                # 2. Find best match in DVKT_GIA_MAX (no chapter filtering for MAX)
                if not matched_by_code_max:
                    match_result_max = self.find_best_match(service_name, max_reference_names, threshold) if max_reference_names else None

                # Logic: Use MAX data if available, otherwise GIA_HDND, otherwise empty
                
                output_row = {
                    'STT': idx + 1,
                    # Default empty
                    'MA_TUONG_DUONG': '',
                    'TEN_DVKT_PHEDUYET': '',
                    'TEN_DVKT_GIA': '',
                    'PHAN_LOAI_PTTT': '',
                    'DON_GIA': '',
                    'GHI_CHU': '',
                    'QUYET_DINH': '',
                    'QUY_TRINH': '',
                    'TU_NGAY': '',
                    'DEN_NGAY': '',
                    'CSKCB_CGKT': '',
                    'CSKCB_CLS': '',
                    'CANH_BAO': ''  # New column for warnings
                }

                # Initialize match tracking variables
                matched_name_display = 'KHÔNG TÌM THẤY'
                score_display = 0
                
                # Check for ambiguous matches first
                is_ambiguous_gia = match_result_gia and match_result_gia[0] == 'AMBIGUOUS'
                is_ambiguous_max = match_result_max and match_result_max[0] == 'AMBIGUOUS'
                
                if is_ambiguous_gia or is_ambiguous_max:
                    # At least one ambiguous match - use first match but warn
                    all_ambiguous = []
                    if is_ambiguous_gia:
                        all_ambiguous.extend(match_result_gia[1])
                    if is_ambiguous_max:
                        all_ambiguous.extend(match_result_max[1])
                    
                    # Remove duplicates
                    seen = set()
                    unique_ambiguous = []
                    for item in all_ambiguous:
                        if item not in seen:
                            seen.add(item)
                            unique_ambiguous.append(item)
                    
                    # Use first match to fill data - prioritize MAX for TEN_DVKT fields, GIA for price
                    first_match = unique_ambiguous[0]
                    
                    # Try to get TEN_DVKT_PHEDUYET and TEN_DVKT_GIA from MAX first (Priority: MAX > GIA)
                    if is_ambiguous_max and match_result_max[1]:
                        max_first = match_result_max[1][0]  # First MAX match
                        ref_matches_max = self.ref_dvkt_gia_max_df[self.ref_dvkt_gia_max_df[max_ref_name_col] == max_first]
                        if len(ref_matches_max) > 0:
                            ref_row_max = ref_matches_max.iloc[0]
                            output_row['TEN_DVKT_GIA'] = max_first
                            output_row['TEN_DVKT_PHEDUYET'] = ref_row_max.get('TEN_DVKT_PHEDUYET', '')
                            output_row['MA_TUONG_DUONG'] = ref_row_max.get('MA_TUONG_DUONG', '')
                            output_row['PHAN_LOAI_PTTT'] = ref_row_max.get('PHAN_LOAI_PTTT', '')
                    
                    # Try to get DON_GIA from GIA (Priority: GIA > MAX)
                    if is_ambiguous_gia and match_result_gia[1]:
                        gia_first = match_result_gia[1][0]  # First GIA match
                        ref_matches_gia = self.ref_gia_hdnd_df[self.ref_gia_hdnd_df[ref_name_column] == gia_first]
                        if len(ref_matches_gia) > 0:
                            ref_row_gia = ref_matches_gia.iloc[0]
                            gia_price = ref_row_gia.get('Mức giá', '')
                            if not pd.isna(gia_price) and str(gia_price).strip() != '':
                                output_row['DON_GIA'] = gia_price
                            output_row['QUYET_DINH'] = ref_row_gia.get('Quyết định', '')
                            # Fallback for TEN_DVKT fields if not set by MAX
                            if not output_row['TEN_DVKT_GIA']:
                                output_row['TEN_DVKT_GIA'] = gia_first
                            if not output_row['TEN_DVKT_PHEDUYET']:
                                output_row['TEN_DVKT_PHEDUYET'] = ref_row_gia.get('Tên chương theo TT 23/2024', '')
                    
                    # GHI_CHU: Take from MAX first, then GIA (same priority as TEN_DVKT fields)
                    if is_ambiguous_max and match_result_max[1]:
                        max_first = match_result_max[1][0]
                        ref_matches_max = self.ref_dvkt_gia_max_df[self.ref_dvkt_gia_max_df[max_ref_name_col] == max_first]
                        if len(ref_matches_max) > 0:
                            output_row['GHI_CHU'] = ref_matches_max.iloc[0].get('GHI_CHU', '')
                    if not output_row['GHI_CHU'] and is_ambiguous_gia and match_result_gia[1]:
                        gia_first = match_result_gia[1][0]
                        ref_matches_gia = self.ref_gia_hdnd_df[self.ref_gia_hdnd_df[ref_name_column] == gia_first]
                        if len(ref_matches_gia) > 0:
                            output_row['GHI_CHU'] = ref_matches_gia.iloc[0].get('Ghi chú', '')
                    
                    # Warning in new last column (CANH_BAO)
                    output_row['CANH_BAO'] = f"⚠️ CÓ {len(unique_ambiguous)} DỊCH VỤ TƯƠNG TỰ - XEM KỸ"
                    
                    matched_name_display = f'{first_match} (⚠️ {len(unique_ambiguous)} tương tự)'
                    score_display = 99
                    stats['ambiguous'] += 1
                    
                    # Add to unmatched with details
                    warning_text = f'⚠️ CÓ {len(unique_ambiguous)} DỊCH VỤ TƯƠNG TỰ:\n'
                    for i, match in enumerate(unique_ambiguous[:5], 1):
                        warning_text += f'{i}. {match}\n'
                    if len(unique_ambiguous) > 5:
                        warning_text += f'... và {len(unique_ambiguous) - 5} dịch vụ khác'
                    
                    unmatched_records.append({
                        'STT': idx + 1,
                        'TEN_DICH_VU_GOC': service_name,
                        'GHI_CHU': warning_text
                    })
                    
                    if self.logger:
                        self.logger.logger.warning(f'Ambiguous match for "{service_name}": Using first of {len(unique_ambiguous)} options')
                    if progress_callback:
                        progress_callback(f"Đã xử lý {idx + 1}/{len(input_df)}: Cảnh báo - {service_name[:50]}...")
                
                elif match_result_max and match_result_max[0] != 'AMBIGUOUS':
                    matched_name_max, score_max = match_result_max
                    ref_matches_max = self.ref_dvkt_gia_max_df[self.ref_dvkt_gia_max_df[max_ref_name_col] == matched_name_max]
                    
                    if len(ref_matches_max) > 0:
                        ref_row_max = ref_matches_max.iloc[0]
                        
                        output_row['MA_TUONG_DUONG'] = ref_row_max.get('MA_TUONG_DUONG', '')
                        output_row['TEN_DVKT_PHEDUYET'] = ref_row_max.get('TEN_DVKT_PHEDUYET', '')
                        output_row['TEN_DVKT_GIA'] = matched_name_max
                        output_row['PHAN_LOAI_PTTT'] = ref_row_max.get('PHAN_LOAI_PTTT', '')
                        output_row['GHI_CHU'] = ref_row_max.get('GHI_CHU', '')
                        
                        max_price = ref_row_max.get('DON_GIA', '')
                        if not max_price and 'GIÁ_MAX' in ref_row_max: 
                             max_price = ref_row_max['GIÁ_MAX']
                        output_row['DON_GIA'] = max_price
                        
                        matched_name_display = matched_name_max
                        score_display = score_max

                # If we have a GIA match (not ambiguous), get Price and other fields
                if match_result_gia and match_result_gia[0] != 'AMBIGUOUS':
                    matched_name_gia, score_gia = match_result_gia
                    ref_matches_gia = self.ref_gia_hdnd_df[self.ref_gia_hdnd_df[ref_name_column] == matched_name_gia]
                    
                    if len(ref_matches_gia) > 0:
                        ref_row_gia = ref_matches_gia.iloc[0]
                        
                        # Price Priority: GIA_HDND > DVKT_GIA_MAX (as requested)
                        # GIA_HDND always takes priority if it has valid price
                        gia_price = ref_row_gia.get('Mức giá', '')
                        if not pd.isna(gia_price) and str(gia_price).strip() != '':
                            # Always use GIA_HDND price if available (overwrite MAX)
                            output_row['DON_GIA'] = gia_price
                        
                        output_row['QUYET_DINH'] = ref_row_gia.get('Quyết định', '') 
                        output_row['TU_NGAY'] = '' 
                        output_row['DEN_NGAY'] = ''
                        
                        # Fallbacks for other fields if not in MAX
                        if not match_result_max:
                            output_row['MA_TUONG_DUONG'] = ref_row_gia.get('Mã tương đương', '')
                            output_row['PHAN_LOAI_PTTT'] = '' 
                            output_row['GHI_CHU'] = ref_row_gia.get('Ghi chú', '')
                            matched_name_display = matched_name_gia
                            score_display = score_gia
                        
                        # TEN_DVKT_PHEDUYET and TEN_DVKT_GIA: Only use GIA if MAX didn't provide them
                        # Priority: MAX > GIA for these two fields
                        if not output_row['TEN_DVKT_PHEDUYET']:
                            output_row['TEN_DVKT_PHEDUYET'] = ref_row_gia.get('Tên chương theo TT 23/2024', '')
                        if not output_row['TEN_DVKT_GIA']:
                            output_row['TEN_DVKT_GIA'] = matched_name_gia

                # LOOKUP QUY_TRINH FOR ALL MATCHED RECORDS (after both GIA and MAX matching)
                # This should run for ALL records that have MA_TUONG_DUONG or a matched name
                if not output_row.get('QUY_TRINH'):  # Only if not already set
                    # Try to find QUY_TRINH from file 1 using MA_TUONG_DUONG (more accurate)
                    ma_tuong_duong = output_row.get('MA_TUONG_DUONG', '')
                    if ma_tuong_duong and str(ma_tuong_duong).strip() and self.ref_quy_trinh_df is not None:
                        # Direct lookup by code - 100% accurate
                        quy_trinh_matches = self.ref_quy_trinh_df[
                            self.ref_quy_trinh_df['MA_DICH_VU'].astype(str).str.strip() == str(ma_tuong_duong).strip()
                        ]
                        if len(quy_trinh_matches) > 0:
                            output_row['QUY_TRINH'] = quy_trinh_matches.iloc[0].get('QUY_TRINH', '')
                            output_row['CSKCB_CGKT'] = quy_trinh_matches.iloc[0].get('CSKCB_CGKT', '')
                            output_row['CSKCB_CLS'] = quy_trinh_matches.iloc[0].get('CSKCB_CLS', '')
                    
                    # If still no QUY_TRINH, fallback to name-based matching
                    if not output_row.get('QUY_TRINH') and quy_trinh_name_col and self.ref_quy_trinh_df is not None:
                        # Use the best matched name we have
                        name_to_lookup = matched_name_display if matched_name_display != 'KHÔNG TÌM THẤY' else service_name
                        quy_trinh_match = self.find_best_match(name_to_lookup, 
                                                               self.ref_quy_trinh_df[quy_trinh_name_col].dropna().astype(str).tolist(),
                                                               threshold)
                        if quy_trinh_match and quy_trinh_match[0] != 'AMBIGUOUS':
                            quy_trinh_matches = self.ref_quy_trinh_df[self.ref_quy_trinh_df[quy_trinh_name_col] == quy_trinh_match[0]]
                            if len(quy_trinh_matches) > 0:
                                quy_trinh_row = quy_trinh_matches.iloc[0]
                                output_row['QUY_TRINH'] = quy_trinh_row.get('QUY_TRINH', '')
                                if not output_row.get('CSKCB_CGKT'):
                                    output_row['CSKCB_CGKT'] = quy_trinh_row.get('CSKCB_CGKT', '')
                                if not output_row.get('CSKCB_CLS'):
                                    output_row['CSKCB_CLS'] = quy_trinh_row.get('CSKCB_CLS', '')

                # Final record
                if not output_row['TEN_DVKT_GIA']:
                    output_row['TEN_DVKT_GIA'] = service_name
                
                results.append(output_row)
                
                # Count stats - ambiguous matches are already counted separately
                if matched_name_display.startswith('CẢNH BÁO'):
                    # Already counted in ambiguous, skip
                    pass
                elif matched_name_display != 'KHÔNG TÌM THẤY':
                     stats['matched'] += 1
                     if self.logger:
                         self.logger.log_match(service_name, matched_name_display, score_display)
                     if progress_callback:
                         progress_callback(f"Đã xử lý {idx + 1}/{len(input_df)}: Khớp - {service_name[:50]}...")
                else:
                     stats['unmatched'] += 1
                     # Only add to unmatched if not already added by ambiguous logic
                     if not any(r['TEN_DICH_VU_GOC'] == service_name for r in unmatched_records):
                         unmatched_records.append({
                             'STT': idx + 1,
                             'TEN_DICH_VU_GOC': service_name,
                             'GHI_CHU': f'Không tìm thấy tên/giá tương ứng trong các file gốc (Ngưỡng khớp: {threshold}%). Vui lòng kiểm tra lại hoặc thêm thủ công.'
                         })
                     if self.logger:
                         self.logger.log_no_match(service_name, threshold)
                     if progress_callback:
                         progress_callback(f"Đã xử lý {idx + 1}/{len(input_df)}: Không khớp - {service_name[:50]}...")
                     
                stats['match_details'].append({
                    'input': service_name,
                    'matched': matched_name_display,
                    'score': score_display
                })
            
            # Create output DataFrame
            output_df = pd.DataFrame(results)
            
            # Save to Excel with proper encoding
            output_df.to_excel(output_file, index=False, engine='openpyxl')
            
            # Apply yellow highlighting to ambiguous rows
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            import os
            
            wb = load_workbook(output_file)
            ws = wb.active
            
            # Yellow fill for ambiguous warnings
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # Find rows with ambiguous warnings (check last column for "⚠️")
            last_col_idx = ws.max_column
            for row_idx in range(2, ws.max_row + 1):  # Start from 2 to skip header
                last_cell = ws.cell(row=row_idx, column=last_col_idx)
                if last_cell.value and isinstance(last_cell.value, str) and '⚠️' in last_cell.value:
                    # Highlight entire row in yellow
                    for col_idx in range(1, last_col_idx + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
            
            # Format Excel for better readability
            # 1. Bold headers
            for col_idx in range(1, last_col_idx + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 2. Auto-fit column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                adjusted_width = min(length + 2, 50)  # Max width 50
                ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
            # 3. Wrap text for long cells
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            wb.save(output_file)
            
            # Auto-open Excel file
            try:
                os.startfile(output_file)  # Windows only
            except:
                pass  # Silently fail if not Windows or file can't be opened
            
            
            # Export unmatched records if any
            if unmatched_records:
                base_name = os.path.splitext(output_file)[0]
                unmatched_file = f"{base_name}_unmatched.xlsx"
                unmatched_df = pd.DataFrame(unmatched_records)
                unmatched_df.to_excel(unmatched_file, index=False, engine='openpyxl')
                stats['unmatched_file'] = unmatched_file
                if self.logger:
                    self.logger.logger.info(f"Đã xuất {len(unmatched_records)} bản ghi không khớp ra: {unmatched_file}")
            
            # Log processing end
            if self.logger:
                self.logger.log_processing_end(stats)
            
            message = f"Xử lý thành công!\n"
            message += f"Tổng số: {stats['total']}\n"
            message += f"Khớp: {stats['matched']}\n"
            message += f"Không khớp: {stats['unmatched']}"
            
            if self.ref_dvkt_gia_max_df is not None:
                message += f"\n(Đã sử dụng dữ liệu ưu tiên từ DVKT_GIA_MAX)"
            
            if stats.get('unmatched_file'):
                message += f"\n\nĐã xuất các bản ghi không khớp ra file:\n{stats['unmatched_file']}"
            
            return True, message, stats
            
        except Exception as e:
            if self.logger:
                self.logger.log_error(str(e))
            return False, f"Lỗi khi xử lý file: {str(e)}", {}
