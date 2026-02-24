# -*- coding: utf-8 -*-
"""Tạo file mẫu GIA_HDND đặt trong public/data/ để người dùng biết cách thay thế."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os

OUT = os.path.join(os.path.dirname(__file__),
                   'web-app', 'public', 'data', 'MAU_GIA_HDND.xlsx')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'GIA_HDND'

# ─── Header ─────────────────────────────────────────────────────────────────
headers = [
    ('Mã tương đương',       '375623'),   # green  = ưu tiên (code match)
    ('Tên dịch vụ kỹ thuật', '1F4E79'),   # blue   = bắt buộc
    ('Mức giá',               '7B3F00'),   # brown  = bắt buộc (giá)
]
widths = [20, 55, 14]
for col, (hdr, color) in enumerate(headers, 1):
    cell = ws.cell(1, col, hdr)
    cell.fill  = PatternFill('solid', fgColor=color)
    cell.font  = Font(bold=True, color='FFFFFF', size=11)
    cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions[ws.cell(1, col).column_letter].width = widths[col - 1]

ws.freeze_panes = 'A2'

# ─── Dữ liệu mẫu ─────────────────────────────────────────────────────────────
rows = [
    ('18.0030.0001', 'Siêu âm tim qua thành ngực',         649000),
    ('09.0012.0003', 'Chụp X-quang phổi thẳng (1 phim)',    53000),
    ('13.0001.0001', 'Xét nghiệm công thức máu',            21000),
    ('17.0055.0002', 'Nội soi dạ dày tá tràng',            350000),
    ('01.0001.0001', 'Khám bệnh (nội khoa)',                 38000),
    ('',             'Ví dụ không có mã — dùng fuzzy tên',  50000),
]
for r, (ma, ten, gia) in enumerate(rows, 2):
    ws.cell(r, 1, ma)
    ws.cell(r, 2, ten)
    ws.cell(r, 3, gia)

# ─── Sheet hướng dẫn ─────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Hướng dẫn')
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 85
HDR = PatternFill('solid', fgColor='1A3A5C')
hdrs2 = ['Cột', 'Loại', 'Mô tả']
for c, h in enumerate(hdrs2, 1):
    cell = ws2.cell(1, c, h)
    cell.fill = HDR
    cell.font = Font(bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center')

guide = [
    ('Mã tương đương',       '⭐ Ưu tiên',
     'Mã XX.XXXX.XXXX — nếu có, khớp chính xác 100%, bỏ qua fuzzy. '
     'Tên cột hợp lệ: Mã tương đương, MA_TUONG_DUONG, MA_DICH_VU, MA_DVKT.'),
    ('Tên dịch vụ kỹ thuật', '✅ Bắt buộc',
     'Tên chuẩn hóa. Dùng fuzzy matching khi không có mã. '
     'Tên cột linh hoạt.'),
    ('Mức giá',               '✅ Bắt buộc',
     'Đơn giá (số). Tên cột hợp lệ: Mức giá, DON_GIA, Giá, Đơn giá.'),
    ('',                     '',
     ''),
    ('Tên file khi thay',    'Lưu ý',
     'Đặt tên GIA_HDND.xlsx, copy vào web-app/public/data/ rồi deploy lại.'),
    ('Upload tạm thời',      'Lưu ý',
     'Hoặc dùng nút 🔄 trong giao diện web để upload ngay, áp dụng cho session đó.'),
]
for r, (a, b, c) in enumerate(guide, 2):
    ws2.cell(r, 1, a)
    ws2.cell(r, 2, b)
    ws2.cell(r, 3, c)

wb.save(OUT)
print(f'✅ Đã tạo: {OUT}')
