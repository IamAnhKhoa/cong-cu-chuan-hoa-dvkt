# -*- coding: utf-8 -*-
"""
Technical Catalog Standardization Tool
Main Application with PyQt5 GUI

Features:
- Easy file/folder selection
- Process type selection
- Fuzzy matching with configurable threshold
- Progress tracking and results preview
"""

import sys
import os
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QPushButton, QLineEdit, 
                             QFileDialog, QProgressBar, QTextEdit, QGroupBox,
                             QRadioButton, QSpinBox, QTableWidget, QTableWidgetItem,
                             QMessageBox, QSplitter, QDialog, QScrollArea)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSettings
from PyQt6.QtGui import QFont
from processor import CatalogProcessor
from ui_styles import MAIN_STYLE
from logger_config import AppLogger


class ProcessingThread(QThread):
    """Background thread for processing files"""
    progress = pyqtSignal(str)
    finished = pyqtSignal(bool, str, dict)
    
    def __init__(self, processor, process_type, input_file, output_file, threshold):
        super().__init__()
        self.processor = processor
        self.process_type = process_type
        self.input_file = input_file
        self.output_file = output_file
        self.threshold = threshold
    
    def run(self):
        """Execute processing in background"""
        self.progress.emit("Đang xử lý file...")
        try:
            if self.process_type == 1:
                success, message, stats = self.processor.process_quy_trinh_file(
                    self.input_file, self.output_file, self.threshold, self.progress.emit
                )
            else:  # process_type == 2
                success, message, stats = self.processor.process_gia_hdnd_file(
                    self.input_file, self.output_file, self.threshold, self.progress.emit
                )
            
            self.finished.emit(success, message, stats)
        except Exception as e:
            self.finished.emit(False, str(e), {})


class ProgressDialog(QDialog):
    """Progress dialog for showing processing status"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Đang xử lý")
        self.setModal(True)
        self.setMinimumWidth(700)
        self.setMinimumHeight(450)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title = QLabel("Đang xử lý file")
        title.setStyleSheet("""
            font-size: 20px; 
            font-weight: bold; 
            color: #2c3e50; 
            padding: 10px 0;
        """)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 0)  # Indeterminate
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: none;
                border-radius: 5px;
                background-color: #ecf0f1;
                height: 8px;
            }
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 5px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        # Status text area
        self.status_text = QTextEdit()
        self.status_text.setReadOnly(True)
        self.status_text.setStyleSheet("""
            QTextEdit {
                background-color: #ffffff;
                border: 1px solid #dfe6e9;
                border-radius: 8px;
                padding: 15px;
                font-family: 'Segoe UI', Arial, sans-serif;
                font-size: 12px;
                color: #2d3436;
                line-height: 1.6;
            }
        """)
        layout.addWidget(self.status_text)
        
        # Bottom info
        info = QLabel("Vui lòng đợi... Cửa sổ sẽ tự động đóng khi hoàn thành")
        info.setStyleSheet("""
            color: #636e72; 
            font-size: 11px; 
            padding: 5px 0;
        """)
        info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info.setWordWrap(True)
        layout.addWidget(info)
        
        self.setLayout(layout)
        
        # Apply overall dialog style
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
        """)
    
    def update_status(self, message):
        """Update status message"""
        self.status_text.append(message)
        # Auto scroll to bottom
        scrollbar = self.status_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
    
    def set_completed(self):
        """Mark as completed"""
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(100)


class MainWindow(QMainWindow):
    """Main application window"""
    
    def __init__(self):
        super().__init__()
        # Initialize logger
        self.app_logger = AppLogger()
        self.app_logger.setup_logger()
        
        self.processor = CatalogProcessor(logger=self.app_logger)
        self.process_type = 1  # Default to process 1
        self.settings = QSettings("MyCompany", "TechnicalCatalogTool")
        self.last_stats = None  # Store last processing stats
        self.init_ui()
        self.load_settings()
        
    def init_ui(self):
        """Initialize user interface"""
        self.setWindowTitle("Công cụ Chuẩn hóa Danh mục Kỹ thuật")
        self.setGeometry(100, 100, 1000, 700)
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15) # Compact spacing
        main_layout.setContentsMargins(20, 20, 20, 20) # Standard margins
        
        # Title
        title = QLabel("CÔNG CỤ CHUẨN HÓA DANH MỤC KỸ THUẬT")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)
        
        # Reference files section
        ref_group = self.create_reference_section()
        main_layout.addWidget(ref_group)
        
        # Process type selection
        process_group = self.create_process_section()
        main_layout.addWidget(process_group)
        
        # Input file section
        input_group = self.create_input_section()
        main_layout.addWidget(input_group)
        
        # Settings section
        settings_group = self.create_settings_section()
        main_layout.addWidget(settings_group)
        
        # Process button
        self.process_btn = QPushButton("BẮT ĐẦU XỬ LÝ")
        self.process_btn.setObjectName("success_button")
        self.process_btn.setMinimumHeight(45) # Slightly smaller
        self.process_btn.clicked.connect(self.start_processing)
        self.process_btn.setEnabled(False)
        main_layout.addWidget(self.process_btn)
        
        # Action buttons
        action_layout = QHBoxLayout()
        action_layout.setSpacing(15)
        self.view_log_btn = QPushButton("Xem nhật ký xử lý")
        self.view_log_btn.setEnabled(False)
        self.view_log_btn.setToolTip("Mở file log để xem chi tiết quá trình xử lý")
        self.view_log_btn.clicked.connect(self.view_log)
        action_layout.addWidget(self.view_log_btn)
        self.view_unmatched_btn = QPushButton("Xem dịch vụ không khớp")
        self.view_unmatched_btn.setEnabled(False)
        self.view_unmatched_btn.setToolTip("Mở file Excel có các dịch vụ không tìm thấy")
        self.view_unmatched_btn.clicked.connect(self.view_unmatched)
        action_layout.addWidget(self.view_unmatched_btn)
        main_layout.addLayout(action_layout)
        
        # Add stretch to prevent widgets from expanding vertically
        main_layout.addStretch()
        
        # Apply styles
        self.setStyleSheet(MAIN_STYLE)
        
    def create_reference_section(self):
        """Create reference files selection section"""
        group = QGroupBox("1. Chọn thư mục chứa file tham chiếu")
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        desc = QLabel('(Chứa: QUY_TRINH, GIA_HDND, DVKT_GIA_MAX)')
        desc.setWordWrap(False)
        desc.setStyleSheet("color: #6c757d; font-size: 12px;")
        layout.addWidget(desc)
        
        folder_layout = QHBoxLayout()
        folder_layout.setSpacing(10)
        self.ref_folder_input = QLineEdit()
        self.ref_folder_input.setPlaceholderText("Chưa chọn thư mục...")
        self.ref_folder_input.setReadOnly(True)
        folder_layout.addWidget(self.ref_folder_input)
        
        browse_btn = QPushButton("Chọn thư mục")
        browse_btn.setMinimumWidth(120)
        browse_btn.clicked.connect(self.browse_reference_folder)
        folder_layout.addWidget(browse_btn)
        
        layout.addLayout(folder_layout)
        
        # Download Sample Button (Moved here)
        download_layout = QHBoxLayout()
        download_layout.addStretch()
        
        download_sample_btn = QPushButton("Tải file mẫu GIA_HDND (399/NQ-HĐND)")
        download_sample_btn.setMinimumWidth(220)
        download_sample_btn.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #138496;
            }
        """)
        download_sample_btn.setToolTip("Tải file Excel mẫu với cấu trúc đúng để import (399/NQ-HĐND)")
        download_sample_btn.clicked.connect(self.download_sample_file)
        download_layout.addWidget(download_sample_btn)
        layout.addLayout(download_layout)
        
        # Status label for loaded files
        self.ref_status = QLabel("")
        self.ref_status.setStyleSheet("color: #28a745; font-size: 12px; font-weight: 600; padding: 5px 0;")
        self.ref_status.setWordWrap(True)
        layout.addWidget(self.ref_status)
        
        group.setLayout(layout)
        return group
    
    def create_process_section(self):
        """Create process type selection section"""
        group = QGroupBox("2. Chọn kiểu xử lý")
        layout = QHBoxLayout() # Horizontal
        layout.setSpacing(15)
        
        self.process1_radio = QRadioButton("Xử lý luồng 1 (TT21 của cổng)")
        self.process1_radio.setChecked(True)
        self.process1_radio.toggled.connect(lambda: self.set_process_type(1))
        layout.addWidget(self.process1_radio)
        
        self.process2_radio = QRadioButton("Xử lý luồng 2 (Import giá dịch vụ được phê duyệt)")
        self.process2_radio.toggled.connect(lambda: self.set_process_type(2))
        layout.addWidget(self.process2_radio)
        
        layout.addStretch()
        
        # Download button removed from here
        
        group.setLayout(layout)
        return group
    
    def create_input_section(self):
        """Create input file selection section"""
        group = QGroupBox("3. Chọn file cần xử lý")
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        desc = QLabel('(Phải có cột: "Tên kỹ thuật" hoặc "Tên dịch vụ")')
        desc.setWordWrap(False)
        desc.setStyleSheet("color: #6c757d; font-size: 12px;")
        layout.addWidget(desc) # Added this line back
        
        file_layout = QHBoxLayout()
        file_layout.setSpacing(10)
        self.input_file = QLineEdit()
        self.input_file.setPlaceholderText("Chưa chọn file...")
        self.input_file.setReadOnly(True)
        file_layout.addWidget(self.input_file)
        
        browse_input_btn = QPushButton("Chọn file")
        browse_input_btn.setMinimumWidth(150)
        browse_input_btn.clicked.connect(self.browse_input_file)
        file_layout.addWidget(browse_input_btn)
        
        layout.addLayout(file_layout)
        
        # Output file
        output_layout = QHBoxLayout()
        output_layout.setSpacing(10)
        
        self.output_file = QLineEdit()
        self.output_file.setPlaceholderText("Tự động đặt tên file kết quả...")
        self.output_file.setReadOnly(True)
        output_layout.addWidget(self.output_file)
        
        output_btn = QPushButton("Chọn vị trí lưu")
        output_btn.setMinimumWidth(150)
        output_btn.clicked.connect(self.browse_output_file)
        output_layout.addWidget(output_btn)
        
        layout.addLayout(output_layout)
        
        group.setLayout(layout)
        return group
    
    def create_settings_section(self):
        """Create settings section"""
        group = QGroupBox("4. Cài đặt nâng cao")
        layout = QHBoxLayout()
        layout.setSpacing(15)
        
        label = QLabel("Ngưỡng khớp tên (%):") 
        layout.addWidget(label)
        
        self.threshold_spin = QSpinBox()
        self.threshold_spin.setRange(50, 100)
        self.threshold_spin.setValue(80)
        self.threshold_spin.setToolTip("Độ chính xác tối thiểu khi so khớp tên dịch vụ (khuyến nghị: 80%)")
        layout.addWidget(self.threshold_spin)
        
        layout.addStretch()
        
        group.setLayout(layout)
        return group
    
    def create_progress_section(self):
        """Create progress and results section"""
        group = QGroupBox("5. Kết quả")
        layout = QVBoxLayout()
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)
        
        # Status text
        self.status_text = QTextEdit()
        self.status_text.setReadOnly(True)
        self.status_text.setMaximumHeight(150)
        self.status_text.setPlaceholderText("Chờ xử lý...")
        layout.addWidget(self.status_text)
        
        group.setLayout(layout)
        return group
    
    def set_process_type(self, process_type):
        """Set the processing type"""
        self.process_type = process_type
        self.update_process_button_state()
    
    def load_settings(self):
        """Load saved settings"""
        saved_folder = self.settings.value("reference_folder")
        if saved_folder and os.path.exists(saved_folder):
            self.ref_folder_input.setText(saved_folder)
            success, message = self.processor.load_reference_files(saved_folder)
            if success:
                self.ref_status.setText(f"✓ {message}")
                self.ref_status.setStyleSheet("color: #28a745; font-size: 12px; font-weight: 600; padding: 5px 0;")
            else:
                self.ref_status.setText(f"✗ Chưa load được file tham chiếu")
                self.ref_status.setStyleSheet("color: #dc3545; font-size: 12px; font-weight: 600; padding: 5px 0;")
            self.update_process_button_state()

    def browse_reference_folder(self):
        """Browse for reference files folder"""
        folder = QFileDialog.getExistingDirectory(
            self, 
            "Chọn thư mục chứa file tham chiếu",
            self.ref_folder_input.text() or os.path.expanduser("~")
        )
        
        if folder:
            self.ref_folder_input.setText(folder)
            self.settings.setValue("reference_folder", folder)
            success, message = self.processor.load_reference_files(folder)
            
            if success:
                # Extract numbers from message (format: "Đã tải X dịch vụ từ QUY_TRINH, Y dịch vụ từ GIA_HDND")
                self.ref_status.setText(f"✓ {message}")
                self.ref_status.setStyleSheet("color: #28a745; font-size: 12px; font-weight: 600; padding: 5px 0;")
            else:
                self.ref_status.setText(f"✗ {message}")
                self.ref_status.setStyleSheet("color: #dc3545; font-size: 12px; font-weight: 600; padding: 5px 0;")
                QMessageBox.warning(self, "Lỗi", message)
            
            self.update_process_button_state()
    
    def browse_input_file(self):
        """Browse for input Excel file"""
        file, _ = QFileDialog.getOpenFileName(
            self,
            "Chọn file Excel đầu vào",
            os.path.expanduser("~"),
            "Excel Files (*.xlsx *.xls)"
        )
        
        if file:
            self.input_file.setText(file)
            # Auto-generate output filename
            if not self.output_file.text():
                base_dir = os.path.dirname(file)
                base_name = os.path.splitext(os.path.basename(file))[0]
                if self.process_type == 1:
                    output_name = f"{base_name}_luong_1.xlsx"
                else:
                    output_name = f"{base_name}_luong_2.xlsx"
                self.output_file.setText(os.path.join(base_dir, output_name))
            
            self.update_process_button_state()
    
    def browse_output_file(self):
        """Browse for output Excel file location"""
        file, _ = QFileDialog.getSaveFileName(
            self,
            "Chọn vị trí lưu file kết quả",
            self.output_file.text() or os.path.expanduser("~"),
            "Excel Files (*.xlsx)"
        )
        
        if file:
            if not file.endswith('.xlsx'):
                file += '.xlsx'
            self.output_file.setText(file)
    
    def update_process_button_state(self):
        """Enable/disable process button based on input state"""
        has_reference = self.processor.ref_quy_trinh_df is not None
        has_input = bool(self.input_file.text())
        has_output = bool(self.output_file.text())
        
        self.process_btn.setEnabled(has_reference and has_input and has_output)
    
    def start_processing(self):
        """Start the processing in background thread"""
        # Disable controls
        self.process_btn.setEnabled(False)
        
        # Create and show progress dialog
        self.progress_dialog = ProgressDialog(self)
        self.progress_dialog.show()
        
        # Create and start processing thread
        self.thread = ProcessingThread(
            self.processor,
            self.process_type,
            self.input_file.text(),
            self.output_file.text(),
            self.threshold_spin.value()
        )
        self.thread.progress.connect(self.progress_dialog.update_status)
        self.thread.finished.connect(self.processing_finished)
        self.thread.start()
    
    def update_progress(self, message):
        """Update progress display - deprecated but kept for compatibility"""
        pass
    
    def processing_finished(self, success, message, stats):
        """Handle processing completion"""
        # Close progress dialog
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.set_completed()
            self.progress_dialog.close()
        
        if success:
            # Show success dialog
            msg_text = f"Xử lý thành công!\n\n{message}\n\nFile kết quả:\n{self.output_file.text()}"
            if stats and stats.get('ambiguous', 0) > 0:
                msg_text += f"\n\n⚠️ Có {stats['ambiguous']} dịch vụ tương tự cần xem kỹ (tô vàng)"
            
            QMessageBox.information(self, "Hoàn thành", msg_text)
        else:
            QMessageBox.critical(self, "Lỗi", f"Xử lý thất bại!\n\n{message}")
        
        # Re-enable controls and store file paths
        self.last_stats = stats  # Store stats for later use
        self.process_btn.setEnabled(True)
        
        # Enable view log button
        if self.app_logger and self.app_logger.get_log_file_path():
            self.view_log_btn.setEnabled(True)
        
        # Enable view unmatched button if unmatched file exists
        if stats and stats.get('unmatched_file'):
            self.view_unmatched_btn.setEnabled(True)
    
    def view_log(self):
        """Open the log file"""
        if self.app_logger and self.app_logger.get_log_file_path():
            log_file = self.app_logger.get_log_file_path()
            if os.path.exists(log_file):
                os.startfile(log_file)
            else:
                QMessageBox.warning(self, "Không tìm thấy", f"File log không tồn tại:\n{log_file}")
    
    def view_unmatched(self):
        """Open the unmatched records file"""
        if self.last_stats and self.last_stats.get('unmatched_file'):
            unmatched_file = self.last_stats.get('unmatched_file')
            if os.path.exists(unmatched_file):
                os.startfile(unmatched_file)
            else:
                QMessageBox.warning(self, "Không tìm thấy", f"File không tồn tại:\n{unmatched_file}")
    
    def create_sample_template(self, file_path):
        """Create sample Excel template for GIA_HDND import"""
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # 1. Create Instruction Sheet Data Structure
        instruction_data = [
            [1, 'Mã tương đương', 'Text', 50, 'Mã dùng để khớp dữ liệu chính xác 100%', 'Khuyến nghị'],
            [2, 'STT', 'Number', 10, 'Số thứ tự', 'Không'],
            [3, 'Tên chương theo TT 23/2024', 'Text', 255, 'Tên chương để lọc và phân loại (Tăng độ chính xác)', 'CÓ'],
            [4, 'Mã kỹ thuật (TT23/2024)', 'Text', 50, 'Mã kỹ thuật theo thông tư mới', 'Không'],
            [5, 'Tên dịch vụ kỹ thuật (TT23/2024)', 'Text', 500, 'Tên dịch vụ theo thông tư mới', 'Không'],
            [6, 'Tên dịch vụ kỹ thuật cụ thể', 'Text', 500, 'Tên dịch vụ thực tế dùng để so khớp (Matching)', 'CÓ'],
            [7, 'Chi phí trực tiếp + Phụ cấp', 'Number', 20, 'Giá trị chi phí', 'Không'],
            [8, 'Lương 1,8 triệu', 'Number', 20, 'Giá trị lương cơ sở cũ', 'Không'],
            [9, 'Giá cụ thể... (TT 21/22)', 'Number', 20, 'Tổng giá theo lương cũ', 'Không'],
            [10, 'Tiền lương 2,34 triệu', 'Number', 20, 'Giá trị lương cơ sở mới', 'Không'],
            [11, 'Giá cụ thể... (lương 2.34)', 'Number', 20, 'Tổng giá theo lương mới', 'Không'],
            [12, 'Mức giá', 'Number', 20, 'Mức giá áp dụng hiện tại (Code ƯU TIÊN lấy giá này)', 'CÓ'],
            [13, 'Ghi chú', 'Text', 255, 'Ghi chú thêm', 'Không'],
            [14, 'Quyết định', 'Text', 50, 'Số quyết định ban hành (Code lấy trực tiếp)', 'CÓ']
        ]
        
        df_instr = pd.DataFrame(instruction_data, columns=['TT', 'Chỉ tiêu', 'Định dạng', 'Kích thước tối đa', 'Diễn giải', 'Bắt buộc'])
        
        # 2. Create Sample Data Sheet
        data = {
            'Mã tương đương': ['01.0002.1778', '01.0004.0321', '01.0006.0215'],
            'STT': [1, 2, 3],
            'Tên chương theo TT 23/2024': [
                '01. HỒI SỨC CẤP CỨU VÀ CHỐNG ĐỘC', 
                '01. HỒI SỨC CẤP CỨU VÀ CHỐNG ĐỘC', 
                '01. HỒI SỨC CẤP CỨU VÀ CHỐNG ĐỘC'
            ],
            'Mã kỹ thuật (TT23/2024)': ['1.2', '1.4', '1.6'],
            'Tên dịch vụ kỹ thuật (TT23/2024)': [
                'Ghi điện tim cấp cứu tại giường', 
                'Ghi điện tim qua chuyển đạo thực quản', 
                'Đặt catheter tĩnh mạch ngoại biên'
            ],
            'Tên dịch vụ kỹ thuật cụ thể': [
                'Ghi điện tim cấp cứu tại giường', 
                'Ghi điện tim qua chuyển đạo thực quản', 
                'Đặt catheter tĩnh mạch ngoại biên'
            ],
            'Chi phí trực tiếp + Phụ cấp': ['20.359', '124.000', '15.000'],
            'Lương 1,8 triệu': ['15.090', '46.957', '7.826'],
            'Giá cụ thể bao gồm chi phí trực tiếp, tiền lương tại TT 21-22': ['35.400', '170.000', '22.800'],
            'Tiền lương 2,34 triệu': ['19.617', '61.043', '10.174'],
            'Giá cụ thể bao gồm chi phí trực tiếp, tiền lương 2,34 trđ': ['39.976', '185.043', '25.174'],
            'Mức giá': ['39.900', '185.000', '25.100'],
            'Ghi chú': [
                '', 
                '', 
                'Chỉ áp dụng với người bệnh ngoại trú; chưa bao gồm thuốc và dịch truyền.'
            ]
        }
        df_data = pd.DataFrame(data)
        
        # Write to Excel with updated sheet order: Sample Data FIRST, Instructions SECOND
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_data.to_excel(writer, sheet_name='Mẫu Import', index=False)
            df_instr.to_excel(writer, sheet_name='Hướng dẫn', index=False)
            
        # Format the Excel file
        wb = load_workbook(file_path)
        
        # Helper function to style headers
        def style_header(ws):
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=10)
            border_style = Side(style='thin', color='000000')
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
                
        # Helper function to auto-fit columns
        def autofit_columns(ws):
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                header_val = column[0].value
                if header_val:
                    max_length = len(str(header_val))
                
                for i, cell in enumerate(column[1:], 1):
                    if i > 20: break
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 4, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Format Data Sheet
        style_header(wb['Mẫu Import'])
        autofit_columns(wb['Mẫu Import'])

        # Format Instruction Sheet
        ws_instr = wb['Hướng dẫn']
        style_header(ws_instr)
        autofit_columns(ws_instr)
        
        # Center align the "Bắt buộc" column
        for cell in ws_instr['F']:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        wb.save(file_path)
    
    def download_sample_file(self):
        """Handle download sample file button click"""
        try:
            # Ask user where to save
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Lưu file mẫu",
                os.path.join(os.path.expanduser("~"), "mau_import_gia_hdnd.xlsx"),
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                # Ensure .xlsx extension
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                
                # Create the sample template
                self.create_sample_template(file_path)
                
                # Show success message
                reply = QMessageBox.question(
                    self,
                    "Thành công",
                    f"Đã tải file mẫu thành công!\n\nVị trí: {file_path}\n\nBạn có muốn mở file ngay không?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                
                if reply == QMessageBox.StandardButton.Yes:
                    os.startfile(file_path)
                    
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể tạo file mẫu:\n{str(e)}")



def main():
    """Main application entry point"""
    app = QApplication(sys.argv)
    
    # Set application-wide font with Vietnamese Unicode support
    # Using Arial which has excellent Vietnamese character support
    font = QFont("Arial", 10)
    font.setStyleHint(QFont.StyleHint.SansSerif)
    app.setFont(font)
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
